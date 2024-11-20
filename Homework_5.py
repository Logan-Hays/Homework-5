#Logan Hays 
#11/19/24
#Homework 5
#Sources used: Chat gtp to help fix some of the errors with colors


import openpyxl

from openpyxl.styles import PatternFill
import string
wb = openpyxl.Workbook()
sheet = wb.active
fill_bla = PatternFill(patternType= 'solid', fgColor= "000000")
fill_ora = PatternFill(patternType= 'solid', fgColor= "ff8000")
fill_yel = PatternFill(patternType= 'solid', fgColor= "fae505")
fill_whi = PatternFill(patternType= 'solid', fgColor= "ffffff")
fill_blu = PatternFill(patternType= 'solid', fgColor= "00f2ff")

cell_bla = ['1D', '1N', '1O', '1P', '1Q', '2C', '2E', '2M', '2R', '3B', '4B', '5A', '6A', '7A', '8B', '8C', '9C', '10C', '11D', '12D', '3E', '4E', '8E', '9E', '13E', '5F', '6F', '7F', '10F', '14F', '11G', '12G', '13G', '14G', '15G', '10H', '11H', '15H', '16H', '17H', '18H', '9I', '16I', '18I', '7J', '8J', '18J', '5K', '6K', '13K', '16K', '18K', '3L', '4L', '11L', '13L', '16L', '17L', '12M', '16M', '15N', '6O', '7O', '8O', '11O', '14O', '15O', '7P', '8P', '11P', '12P', '13P', '15P', '11Q', '14Q', '10R', '3S', '4S', '10S', '5T', '9T', '6U', '7U', '8U']
cell_o = ['9D', '10D', '10E', '11E', '12E', '11F', '12F', '13F', '12H', '13H', '14H', '10I','11I', '12I', '13I', '14I', '15I', '9J','10J', '11J', '12J', '13J', '14J', '15J', '16J', '17J', '7K', '8K', '9K', '10K', '11K', '12K', '14K', '15K', '5L', '6L', '7L', '8L', '9L', '10L', '12L', '3M', '4M', '5M', '6M', '7M', '8M', '9M', '10M', '11M', '2N', '3N', '4N', '5N', '6N', '7N', '8N', '9N', '10N', '11N', '2O', '3O', '4O', '5O', '9O', '10O', '2P', '3P', '4P', '5P', '9P', '10P', '2Q', '3Q', '4Q', '5Q', '6Q', '7Q', '8Q', '9Q', '10Q', '3R', '4R', '5R', '6R', '7R', '8R', '9R', '5S', '6S', '7S', '8S', '9S', '6T', '7T', '8T']
cell_y = ['6C', '7C', '7D', '8D', '14L', '15L', '13M', '14M', '15M', '12N', '13N', '14N', '12O', '13O']
cell_w = ['17I', '17K', '6P']

for chr in string.ascii_uppercase[:22]:
    sheet.column_dimensions[chr].width = 5
    for i in range (1, 20):
        sheet.row_dimensions[i].height = 20
        coord = chr + str(i)
        if coord in cell_bla:
            sheet[coord].fill = fill_bla
        elif coord in cell_o:
            sheet[coord].fill = fill_ora
        elif coord in cell_y:
            sheet[coord].fill = fill_yel
        elif coord in cell_w:
            sheet[coord].fill = fill_whi
        else:
            sheet[coord].fill = fill_blu

wb.save("Homework_5.xlsx")