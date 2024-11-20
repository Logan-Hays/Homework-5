#Logan Hays 
#Lab section: 13
#11/19/24
#Homework 5
#Sources used: Chat gtp to help fix some of the errors with colors


import openpyxl

from openpyxl.styles import PatternFill
wb = openpyxl.Workbook()
sheet = wb.active
fill_bla = PatternFill(patternType= 'solid', fgColor= "000000")
fill_ora = PatternFill(patternType= 'solid', fgColor= "ff8000")
fill_yel = PatternFill(patternType= 'solid', fgColor= "fae505")
fill_whi = PatternFill(patternType= 'solid', fgColor= "ffffff")
fill_red = PatternFill(patternType= 'solid', fgColor= "ff0000")

for row in range(1, 20):
    sheet.row_dimensions[row].height = 20
for column in range(1, 22):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 5

cell_bla = ['D1', 'N1', 'O1', 'P1', 'Q1', 'C2', 'E2', 'M2', 'R2', 'B3', 'B4', 'A5', 'A6', 'A7', 'B8', 'C8', 'C9', 'C10', 'D11', 'D12', 'E3', 'E4', 'E8', 'E9', 'E13', 'F5', 'F6', 'F7', 'F10', 'F14', 'G11', 'G12', 'G13', 'G14', 'G15', 'H10', 'H11', 'H15', 'H16', 'H17', 'H18', 'I9', 'I16', 'I18', 'J7', 'J8', 'J18', 'K5', 'K6', 'K13', 'K16', 'K18', 'L3', 'L4', 'L11', 'L13', 'L16', 'L17', 'M12', 'M16', 'N15', 'O6', 'O7', 'O8', 'O11', 'O14', 'O15', 'P7', 'P8', 'P11', 'P12', 'P13', 'P15', 'Q11', 'Q14', 'R10', 'S3', 'S4', 'S10', 'T5', 'T9', 'U6', 'U7', 'U8']
for cell in cell_bla:
    sheet[cell].fill = fill_bla
cell_o = ['D9', 'D10', 'E10', 'E11', 'E12', 'F11', 'F12', 'F13', 'H12', 'H13', 'H14', 'I10', 'I11', 'I12', 'I13', 'I14', 'I15', 'J9', 'J10', 'J11', 'J12', 'J13', 'J14', 'J15', 'J16', 'J17', 'K7', 'K8', 'K9', 'K10', 'K11', 'K12', 'K14', 'K15', 'L5', 'L6', 'L7', 'L8', 'L9', 'L10', 'L12', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8', 'M9', 'M10', 'M11', 'N2', 'N3', 'N4', 'N5', 'N6', 'N7', 'N8', 'N9', 'N10', 'N11', 'O2', 'O3', 'O4', 'O5', 'O9', 'O10', 'P2', 'P3', 'P4', 'P5', 'P9', 'P10', 'Q2', 'Q3', 'Q4', 'Q5', 'Q6', 'Q7', 'Q8', 'Q9', 'Q10', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'S5', 'S6', 'S7', 'S8', 'S9', 'T6', 'T7', 'T8']
for cell in cell_o:
    sheet[cell].fill = fill_ora
cell_y = ['C6', 'C7', 'D7', 'D8', 'L14', 'L15', 'M13', 'M14', 'M15', 'N12', 'N13', 'N14', 'O12', 'O13']
for cell in cell_y:
    sheet[cell].fill = fill_yel
cell_w = ['I17', 'K17', 'P6']
for cell in cell_w:
    sheet[cell].fill= fill_whi
cell_r = ['B5', 'B6', 'B7', 'C3', 'C4', 'C5', 'D2', 'D3', 'D4', 'D5', 'D6', 'E5', 'E6', 'E7']
for cell in cell_r:
    sheet[cell].fill= fill_red


wb.save("Homework_5.xlsx")